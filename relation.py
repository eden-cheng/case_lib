import yaml
import openpyxl
import openpyxl
from ruamel import yaml
import re

file_name = 'para.yaml'
yaml_relation_file = "./yaml_files/relation.yaml"

def yaml_manage(file_path):
    """载入yaml配置文件，并存为字典格式"""
    with open (file_path, encoding = 'utf-8') as file_obj:
        content = file_obj.read()
        yaml_dic = yaml.load(content, Loader=yaml.Loader)
        return yaml_dic

def func_yaml():
    with open('./yaml_files/relation.yaml', encoding='utf-8') as obj:
        content = obj.read()
        dic = yaml.load(content, Loader=yaml.Loader)    #ruamel中的读取，不是FullLoader，而是Loader
        return dic

def func_para(ws, num):
    num_int = int(num)
    return ws[num_int]

# 将下面形式字符串转换成列表结构
# a;
# b;
# c;
def func_odd_1(string):
    odd_1 = string.split('\n')
    return odd_1

#将a;b;c;转换成列表结构
def func_odd_2(string):
    odd_2 = string.split(' ')
    return odd_2

#将 A：B 字符串转成字典结构
def func_dic(string):
    dic = {}
    res = string.split(':')
    dic[res[0]] = res[1]
    return dic

#将下面形式转化成列表和字典嵌套格式
# A:1;2;3;#
# B:1;2;#
# C:0;1;2;3;4;#
def func(string):
    arr_1 = func_odd_1(string)
    arr_2 = []
    for i in arr_1:
        dic = {}
        for key,value in func_dic(i).items():
            arr = func_odd_2(value)
            dic[key] = arr
            arr_2.append(dic)
    return arr_2

#写入yaml文件
def output_yaml(arr, file_name):
    with open(file_name, 'a+', encoding='utf-8') as f:
        yaml.dump(arr, f, Dumper=yaml.RoundTripDumper)

# 行列关系
def col_relation(ws):
    with open (yaml_relation_file, encoding = 'utf-8') as file_obj:
        content = file_obj.read()
        yaml_dic = yaml.load(content, Loader=yaml.Loader)
        case_col = yaml_dic['output_titles']
        # case_row = list(yaml_dic['row_relation'].values())
    dic_col = {}
    for columns in ws[2]:
        for col_name in case_col:
            if col_name == columns.value:
                dic_col[col_name] = columns.column
    return dic_col
def row_relation(ws):
    dic_row = {}
    for rows in ws['A']:
        list_row = []
        list_row.append(rows.value)
        reg = re.compile('\w{6,8}')
        for i in list_row:
            res = reg.search(str(i))
            if res:
                # print(res.group())
                if res.group() == rows.value:
                    dic_row[res.group()] = rows.row
    return dic_row

# 循环五行
def para_to_yaml():
    res = {}
    temp = []
    dic_row = row_relation(ws_para)
    for v in dic_row.values():
        temp.append(v)
    for i in temp:
        dic_g = {}
        for j in (i, i+5):
            cell_action = ws_para.cell(row=j, column=3).value
            cell_odd = ws_para.cell(row=j, column=4).value
            dic_action = func(cell_action)
            dic_odd = func(cell_odd)
            print(dic_action)
            # 转化成字典
            # if dic_action or dic_odd:

            #     dic_g['group' + str(j)] = dict(dic_action, **dic_odd)

wb_relation = openpyxl.load_workbook('TC_lib.xlsx')
ws_relation = wb_relation['CC-lib']
# print(col_relation(ws_relation))
# print(row_relation(ws_para))

wb_para = openpyxl.load_workbook("para.xlsx")
ws_para = wb_para['CC-para']

# row_1 = func_para(ws_para, func_yaml()['CC_1_1'])
# row_2 = func_para(ws_para, func_yaml()['CC_2_1'])
# row_3 = func_para(ws_para, func_yaml()['CC_3_1'])

# para_odd_1 = row_1[3].value
# para_odd_2 = row_2[3].value
# para_action_1 = row_2[2].value
# para_action_2 = row_3[2].value

# print(func_odd_1(para_odd_1))
# print(func_odd_2(para_odd_2))

# arr = func(para_action_2)
# arr_col = col_relation(ws_relation)

# lib行号
# print(row_relation(ws_relation))
# arr_row = row_relation(ws_para)
# para行号
# print(row_relation(ws_para))
# arr_row = row_relation(ws_relation)

#将字典和列表的嵌套格式转成yaml方式写入
# output_yaml(arr_row, yaml_relation_file)

para_to_yaml()