import openpyxl
import yaml
import os

yaml_path_file = "./yaml_files/file.yaml"
yaml_relation_file = "./yaml_files/relation.yaml"

class Tool():
    def yaml_manage(self, file_path):
        """载入yaml配置文件，并存为字典格式"""
        with open (file_path, encoding = 'utf-8') as file_obj:
            content = file_obj.read()
            yaml_dic = yaml.load(content, Loader=yaml.FullLoader)
            return yaml_dic

    def charact_value(self, file_path, test_name):
        """将yaml中的多组特征值文件路径，提取出成参数字典"""
        #实例化 Yaml_manage 类，并从yaml中提取出实车测试特征值文件的路径构成字典
        #键是实车测试对象，比如noload, payload
        #值是特征值文件的路径
        yaml_s = (self.yaml_manage(file_path))[test_name]
        #分别将实车测试特征值文件中的内容提取出来，形成一个庞大的特征值字典。
        #键是实车测试的对象， 比如noload, payload
        #值是特征值
        cc_veh_value = {}
        for key, value in yaml_s.items():
            cc_veh_value[key] = self.yaml_manage(value)
        return cc_veh_value

    def import_data(self, arr, yaml_title):
        """将output整理的列表，写入表格，仅做写入，不做列表的追加等操作"""
        out_path = (self.yaml_manage(yaml_path_file))['case_path_cc_veh']
        wb = openpyxl.Workbook()
        ws = wb.active
        output_titles = (self.yaml_manage(yaml_relation_file))['output_titles']
        
        #将标题填入excel 
        for i in range(len(output_titles)):
            ws.cell(row = 1, column = i + 1).value = output_titles[i]
        #将expand整理出的case，从字典形式填入Excel中
        j = 2
        for dic in arr:
            k = 1
            for title, content in dic.items():
                ws.cell(row = j, column = k).value = dic[output_titles[k-1]]
                k += 1
            j = j + 1
        #python生成文件路径
        isExists = os.path.exists(out_path)
        if not isExists:
            os.makedirs(file_path)
        wb.save(out_path + yaml_title + '.xlsx')